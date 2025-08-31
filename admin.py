from flask import Flask, render_template, request, redirect, url_for, send_file, Response
import openpyxl, io, os, base64
import boto3
from botocore.exceptions import ClientError

app = Flask(__name__)

# --- Basic Auth (HTTP) ---
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("ADMIN_PASS", "changeme")

def require_basic_auth():
    def decorator(view):
        def wrapped(*args, **kwargs):
            auth = request.headers.get("Authorization", "")
            if auth.startswith("Basic "):
                try:
                    userpass = base64.b64decode(auth.split()[1]).decode("utf-8")
                except Exception:
                    return _auth_failed()
                parts = userpass.split(":", 1)
                if len(parts) == 2:
                    u, p = parts
                    if u == ADMIN_USER and p == ADMIN_PASS:
                        return view(*args, **kwargs)
            # If not authorized, prompt
            return _auth_failed()
        wrapped.__name__ = view.__name__
        return wrapped
    return decorator

def _auth_failed():
    return Response(
        "Authentication required",
        401,
        {"WWW-Authenticate": 'Basic realm="Admin Area"'}
    )

# --- S3 settings ---
S3_BUCKET = os.getenv("S3_BUCKET")
S3_KEY = os.getenv("S3_KEY", "data.xlsx")
AWS_REGION = os.getenv("AWS_REGION", "us-east-1")

def _s3_client():
    return boto3.client("s3", region_name=AWS_REGION)

def _load_workbook_from_s3():
    if not S3_BUCKET:
        # fallback to local file for dev
        return openpyxl.load_workbook("data.xlsx")
    s3 = _s3_client()
    bio = io.BytesIO()
    s3.download_fileobj(S3_BUCKET, S3_KEY, bio)
    bio.seek(0)
    return openpyxl.load_workbook(bio)

def _save_workbook_to_s3(wb):
    if not S3_BUCKET:
        wb.save("data.xlsx")
        return
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    s3 = _s3_client()
    s3.upload_fileobj(bio, S3_BUCKET, S3_KEY, ExtraArgs={"ContentType":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})

def load_guests():
    wb = _load_workbook_from_s3()
    sh = wb.active
    guests = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        guests.append({"first": row[0], "last": row[1], "nick": row[2], "table": row[3]})
    return guests

def save_guests(guests):
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.append(["FirstName", "LastName", "Nickname", "TableNumber"])
    for g in guests:
        sh.append([g["first"], g["last"], g["nick"], g["table"]])
    _save_workbook_to_s3(wb)

@app.route("/", methods=["GET"])
@require_basic_auth()
def index():
    return render_template("admin.html", guests=load_guests())

@app.route("/edit/<int:index>", methods=["GET", "POST"])
@require_basic_auth()
def edit(index):
    guests = load_guests()
    guest = guests[index]
    if request.method == "POST":
        guest["first"] = request.form["first"]
        guest["last"]  = request.form["last"]
        guest["nick"]  = request.form["nick"]
        guest["table"] = request.form["table"]
        save_guests(guests)
        return redirect(url_for("index"))
    return render_template("edit.html", guest=guest, index=index)

@app.route("/delete/<int:index>", methods=["POST"])
@require_basic_auth()
def delete(index):
    guests = load_guests()
    if 0 <= index < len(guests):
        guests.pop(index)
        save_guests(guests)
    return redirect(url_for("index"))

@app.route("/upload", methods=["POST"])
@require_basic_auth()
def _normalize_header(val):
    return str(val or "").strip().lower()

def _validate_headers(ws):
    expected = ["firstname","lastname","nickname","tablenumber"]
    headers = [ _normalize_header(ws.cell(1, i+1).value) for i in range(4) ]
    return headers == expected

def upload():
    f = request.files.get("file")
    if not f or not f.filename.endswith(".xlsx"):
        return Response("Please upload a .xlsx file", 400)
    wb = openpyxl.load_workbook(f.stream)
    ws = wb.active
    if not _validate_headers(ws):
        return Response("Invalid header row. Expected: FirstName, LastName, Nickname, TableNumber", 400)
    # (Optional) quick sanity: ensure no completely empty required cells in data rows
    # and that TableNumber is provided where names exist.
    for r in ws.iter_rows(min_row=2, values_only=True):
        first, last, nick, table = r[:4]
        if any([first, last, nick, table]):
            if not (first and last and nick and table is not None):
                return Response("Found a row with missing values. Each row must have FirstName, LastName, Nickname, TableNumber.", 400)
    _save_workbook_to_s3(wb)
    return redirect(url_for("index"))

@app.route("/download", methods=["GET"])
@require_basic_auth()
def download():
    # Always stream from S3 so you get the latest
    if not S3_BUCKET:
        return send_file("data.xlsx", as_attachment=True, download_name="guest_list_export.xlsx")
    s3 = _s3_client()
    bio = io.BytesIO()
    s3.download_fileobj(S3_BUCKET, S3_KEY, bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="guest_list_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__ == "__main__":
    app.run(debug=True, port=5001)
