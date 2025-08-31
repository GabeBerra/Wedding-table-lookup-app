from flask import Flask, render_template, request, send_file, url_for
import io, os
import openpyxl
import qrcode
import boto3
from botocore.exceptions import ClientError

app = Flask(__name__)

# S3 settings from environment
S3_BUCKET = os.getenv("S3_BUCKET")
S3_KEY = os.getenv("S3_KEY", "data.xlsx")
AWS_REGION = os.getenv("AWS_REGION", "us-east-1")

def _s3_client():
    return boto3.client("s3", region_name=AWS_REGION)

def _load_sheet_from_s3():
    if not S3_BUCKET:
        # Fallback to local file for dev
        wb = openpyxl.load_workbook("data.xlsx")
        return wb.active

    s3 = _s3_client()
    bio = io.BytesIO()
    try:
        s3.download_fileobj(S3_BUCKET, S3_KEY, bio)
    except ClientError as e:
        # On first run if object missing, fallback to local file if exists
        if os.path.exists("data.xlsx"):
            wb = openpyxl.load_workbook("data.xlsx")
            return wb.active
        raise e
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    return wb.active

def find_table(first, last, nick):
    sh = _load_sheet_from_s3()
    for row in sh.iter_rows(min_row=2, values_only=True):
        if (str(row[0]).strip().lower() == first.strip().lower() and
            str(row[1]).strip().lower() == last.strip().lower() and
            str(row[2]).strip().lower() == nick.strip().lower()):
            return row[3]
    return None

def get_guests_by_table(table_number):
    sh = _load_sheet_from_s3()
    guests = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if str(row[3]) == str(table_number):
            guests.append({"first": row[0], "last": row[1], "nick": row[2]})
    return guests

@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    guests = []
    if request.method == "POST":
        first = request.form["first_name"]
        last = request.form["last_name"]
        nick = request.form["nickname"]
        result = find_table(first, last, nick)
        if result:
            guests = get_guests_by_table(result)
    return render_template("index.html", result=result, guests=guests)

# --- QR: dynamic PNG for the site's URL ---
@app.route("/qr.png")
def qr_png():
    target_url = request.url_root.rstrip("/") + url_for("index")
    img = qrcode.make(target_url)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return send_file(bio, mimetype="image/png", download_name="guest_site_qr.png")

# --- QR landing page for printing ---
@app.route("/qr")
def qr_page():
    return render_template("qr.html")

if __name__ == "__main__":
    app.run(debug=True, port=5000)
